# -*- coding: utf-8 -*-
"""Outstanding-payments sheet -> JSON (LOCAL). Seeds the LIVE collections
worklist (neon.collections.item). Parses the 12 accounts, seeds a sortable
status from the note (note kept VERBATIM as source of truth), splits the
contact name/phone, tags period_year from the section headers, and flags the
3 data quirks. Partner + sales-rep links are resolved by the LOADER (need prod).

Columns (0-indexed): 0 client | 1 event/venue | 2 USD | 3 ZWG | 4 contact |
5 sales rep | 6 note. Section-header rows ("2025/2026 Outstanding Payments")
set period_year (not data). Total row is a sum (skip).

Usage: python parse_collections.py [out.json]
"""
import json
import re
import sys

DEFAULT_XLSX = r"C:\Users\Neon\Downloads\Outstanding payments Neon.xlsx"
SHEET = "Outstanding Payment Records"


def seed_status(note):
    """Map the verbatim note -> a sortable status (note always kept)."""
    n = (note or "").lower()
    if not n.strip():
        return "chasing"                       # blank default
    if "recovered" in n:
        return "recovered"
    if ("promise to pay" in n or "payment plan" in n
            or "will be transfered" in n or "awaiting pop" in n):
        return "promised"
    if "po submitted" in n or "po available" in n:
        return "po_submitted"
    if "80%" in n or "balance after event" in n:
        return "part_paid"
    if "ignoring calls" in n:
        return "unresponsive"
    if "foreign payment" in n or "likely to clear" in n:
        return "clearing"
    if "checking" in n or "balance tba" in n:
        return "chasing"
    return "chasing"


def split_contact(s):
    """'Rati - 0782724481' -> ('Rati', '0782724481'); 'Vusa' -> ('Vusa','')."""
    s = (s or "").strip()
    if not s:
        return "", ""
    m = re.search(r"(0\d[\d\s]{6,})", s)
    phone = re.sub(r"\s+", "", m.group(1)) if m else ""
    name = s
    if m:
        name = s[:m.start()].rstrip(" -––").strip()
    return name, phone


def parse(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    period = None
    for r in rows:
        c0 = (str(r[0]).strip() if len(r) > 0 and r[0] is not None else "")
        if not c0:
            continue
        low = c0.lower()
        if "outstanding payments" in low:           # section header
            period = "2025" if "2025" in c0 else ("2026" if "2026" in c0
                                                  else period)
            continue
        if low in ("total", "totals", "grand total"):
            continue
        if low in ("venue",):
            continue
        event = (str(r[1]).strip() if len(r) > 1 and r[1] is not None else "")
        usd = r[2] if len(r) > 2 else None
        zwg_cell = r[3] if len(r) > 3 else None
        contact = (str(r[4]).strip() if len(r) > 4 and r[4] is not None else "")
        rep = (str(r[5]).strip() if len(r) > 5 and r[5] is not None else "")
        note = (str(r[6]).strip() if len(r) > 6 and r[6] is not None else "")
        amount_usd = float(usd) if isinstance(usd, (int, float)) else None
        amount_zwg = float(zwg_cell) if isinstance(zwg_cell, (int, float)) \
            else None
        currency_flag = ""
        if isinstance(zwg_cell, str) and zwg_cell.strip():
            currency_flag = "%s (verify)" % zwg_cell.strip()  # "ZWG Payment"
        cname, cphone = split_contact(contact)
        out.append({
            "client_name": c0, "event_name": event,
            "amount_usd": amount_usd, "amount_zwg": amount_zwg,
            "currency_flag": currency_flag,
            "contact_name": cname, "contact_phone": cphone,
            "sales_rep_raw": rep, "note": note,
            "status": seed_status(note), "period_year": period or "2026",
            "source": "outstanding_sheet",
        })
    return {"items": out}


def main():
    out = next((a for a in sys.argv[1:] if a.endswith(".json")), None)
    p = parse(DEFAULT_XLSX)
    if out:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(p, f, indent=1, default=str)
    return p


if __name__ == "__main__":
    p = main()
    items = p["items"]
    tot = sum(i["amount_usd"] or 0 for i in items)
    print("COLLECTIONS PARSE: %d items, USD total %.2f" % (len(items), tot))
    print("  %-26s %-20s %-9s %-12s %-10s %s"
          % ("CLIENT", "EVENT", "USD", "STATUS", "REP", "yr"))
    for i in items:
        print("  %-26s %-20s %-9s %-12s %-10s %s | note=%r"
              % (i["client_name"][:26], i["event_name"][:20],
                 i["amount_usd"], i["status"], i["sales_rep_raw"],
                 i["period_year"], i["note"][:40]))
    print("\n  contact split + currency flags:")
    for i in items:
        if i["currency_flag"] or i["contact_phone"]:
            print("    %-20s name=%-10r phone=%-12r flag=%r"
                  % (i["client_name"][:20], i["contact_name"],
                     i["contact_phone"], i["currency_flag"]))
