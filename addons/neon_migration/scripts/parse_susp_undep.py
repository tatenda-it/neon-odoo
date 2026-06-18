# -*- coding: utf-8 -*-
"""Suspense + Undeposited ledgers xlsx -> JSON parser (LOCAL, reference-only).

Stores BOTH ledgers VERBATIM (faithful, like petty cash) — income + transfers +
expenses, not cherry-picked. Inert; NOT account.move.

SUSPENSE (2 tabs: 2025 ' Suspense Account', 2026 '2026 Suspense Acc'):
  6-col running-balance cashbook (Date/Details/Acc/Dr/Cr/Balance). MULTI-MONTH
  (no tab-month anchor) -> the authoritative integrity check is the running-
  balance reconciliation; dates are best-effort (text DD-MM-YY parsed; ambiguous
  datetime cells -> None, NEVER fabricated). period_month = year anchor.

UNDEPOSITED (6 tabs, 2025; monthly so dates use the petty-cash per-cell decode):
  * two_table  (Jan, Feb): receipts (Date/Details/Invoice No/Invoice Amount
    [+ZWG col in Feb][+method/note]) then expenses (Date/Details/Amount/Account).
  * dr_cr      (April): Date/Details/Acc/Dr/Cr.
  * amount     (May, June): Date/Details/Acc/Amount.
  * empty      (July): surfaced as an empty statement (NOT silently dropped).
  Captured verbatim into a flexible line (section / amount / dr / cr / currency /
  invoice_no / note). Feb ZWG lines get currency='ZWG'.

openpyxl is imported lazily in main() so the pure helpers unit-test in the Odoo
shell with synthetic rows.
"""
import calendar
import datetime
import json
import re
import sys

DEFAULT_2025 = r"C:\Users\Neon\Downloads\Neon Expenses 2025.xlsx"
DEFAULT_2026 = r"C:\Users\Neon\Downloads\Neon Expenses 2026.xlsx"

_MONTH_WORDS = {}
for _i, _full in enumerate(calendar.month_name):
    if _full:
        _MONTH_WORDS[_full.lower()] = _i
for _i, _ab in enumerate(calendar.month_abbr):
    if _ab:
        _MONTH_WORDS[_ab.lower()] = _i
_MONTH_WORDS["sept"] = 9
_SUMMARY = ("closing balance", "total", "totals", "grand total")


def _norm(c):
    return str(c).strip().lower() if c is not None else ""


def _month_from_tabname(tab):
    for w in re.findall(r"[a-zA-Z]+", (tab or "").lower()):
        if w in _MONTH_WORDS:
            return _MONTH_WORDS[w]
    return None


def _num(c):
    return float(c) if isinstance(c, (int, float)) else None


def _text_date(cell):
    """text 'DD-MM-YY[YY]' or 'DD/MM/YY' -> ISO date string, else None."""
    if not isinstance(cell, str):
        return None
    m = re.match(r"\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{2,4})",
                 cell.strip())
    if not m:
        return None
    dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    yy = 2000 + yy if yy < 100 else yy
    try:
        return datetime.date(yy, mm, dd).isoformat()
    except ValueError:
        return None


def _decode_day_monthly(cell, tab_month, prev_day):
    """Per-cell day decode for MONTHLY tabs (undeposited). Mirrors petty cash:
    mangled datetime(Y,day,tab_month) / real datetime(Y,tab_month,day) / text
    DD-MM-YY / blank carry. Returns (day, kind)."""
    if cell is None or (isinstance(cell, str) and not cell.strip()):
        return prev_day, "carry"
    if isinstance(cell, datetime.datetime):
        if cell.day == tab_month:
            return cell.month, "dt-mangled"
        if cell.month == tab_month:
            return cell.day, "dt-real"
        return prev_day, "dt-ambiguous"
    if isinstance(cell, str):
        m = re.match(r"\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{2,4})",
                     cell.strip())
        if m:
            return int(m.group(1)), "text"
        return prev_day, "text-note"
    return prev_day, "unparsed"


def _find_rows_like(rows, predicate):
    """Yield (row_index, col_offset) for rows whose 6-col window from some
    offset satisfies predicate(seg)."""
    out = []
    for i, row in enumerate(rows):
        rl = list(row)
        for c in range(0, max(1, len(rl) - 3)):
            seg = [_norm(x) for x in (rl[c:c + 6] + [""] * 6)[:6]]
            if predicate(seg):
                out.append((i, c))
                break
    return out


# ----------------------------------------------------------------------
# SUSPENSE (6-col running balance, multi-month)
# ----------------------------------------------------------------------
def _hdr_suspense(seg):
    return (seg[0].startswith("date") and seg[1].startswith("details")
            and seg[5].startswith("balance"))


def parse_suspense(rows, tab, year):
    hdrs = _find_rows_like(rows, _hdr_suspense)
    if not hdrs:
        return None, ["no suspense header"]
    header_idx, col0 = hdrs[0]
    issues = []
    if col0:
        issues.append("column offset %d" % col0)
    txns = []
    gap = 0
    i = header_idx + 1
    while i < len(rows):
        row = list(rows[i][col0:col0 + 6]) + [None] * 6
        details = row[1]
        dl = _norm(details)
        is_sum = any(s in dl for s in _SUMMARY) if dl else False
        has_any = any(x is not None for x in row[:6])
        if details is not None and dl and not is_sum:
            txns.append(row)
            gap = 0
        else:
            gap += 1
            if txns and gap >= 8:
                break
        i += 1

    lines = []
    prev_bal = None
    sum_dr = sum_cr = 0.0
    recon_fails = []
    for seq, row in enumerate(txns, start=1):
        date_cell, details, acc, dr, cr, bal = row[:6]
        dr = _num(dr) or 0.0
        cr = _num(cr) or 0.0
        bal = _num(bal)
        sum_dr += dr
        sum_cr += cr
        if prev_bal is not None and bal is not None:
            if abs((prev_bal + dr - cr) - bal) > 0.01:
                recon_fails.append((seq, prev_bal, dr, cr, bal))
        if bal is not None:
            prev_bal = bal
        lines.append({
            "sequence": seq * 10,
            "date_raw": (date_cell.isoformat()
                         if isinstance(date_cell, datetime.datetime)
                         else (date_cell if date_cell is not None else "")),
            "date_parsed": _text_date(date_cell),  # datetime -> None (ambig)
            "details": (str(details).strip() if details else ""),
            "acc_code": (str(acc).strip() if acc else ""),
            "debit": dr, "credit": cr, "balance": bal,
        })
    closing_line = lines[-1]["balance"] if lines else None
    return {
        "kind": "suspense", "tab": tab, "year": year,
        "name": "Suspense Account %s" % year,
        "period_month": "%d-01-01" % year,
        "currency_code": "USD",
        "opening_balance": 0.0,  # netting account starts at 0
        "closing_balance": closing_line,
        "line_count": len(lines), "lines": lines,
        "v_sum_dr": round(sum_dr, 2), "v_sum_cr": round(sum_cr, 2),
        "v_recon_fails": recon_fails,
        "v_balance_eq": (closing_line is not None
                         and abs((sum_dr - sum_cr) - closing_line) < 0.01),
        "issues": issues,
    }, issues


# ----------------------------------------------------------------------
# UNDEPOSITED (4 formats)
# ----------------------------------------------------------------------
def _hdr_has(seg, *needles):
    j = " ".join(seg)
    return seg[0].startswith("date") and all(n in j for n in needles)


def parse_undeposited(rows, tab):
    tab_month = _month_from_tabname(tab)
    nonblank = [i for i, r in enumerate(rows) if any(x is not None for x in r)]
    if not nonblank:
        return {"kind": "undeposited", "tab": tab, "tab_month": tab_month,
                "statement_format": "empty", "line_count": 0, "lines": [],
                "issues": ["empty tab"], "totals": {}}, ["empty tab"]

    receipts_hdr = _find_rows_like(rows, lambda s: _hdr_has(s, "invoice"))
    exp_hdr = _find_rows_like(rows, lambda s: _hdr_has(s, "amount", "account"))
    drcr_hdr = _find_rows_like(
        rows, lambda s: s[0].startswith("date")
        and ("dr" in s or "cr" in s) and "balance" not in " ".join(s))
    amount_hdr = _find_rows_like(
        rows, lambda s: _hdr_has(s, "amount") and "account" not in " ".join(s)
        and "invoice" not in " ".join(s))

    if receipts_hdr:
        fmt = "two_table"
    elif drcr_hdr:
        fmt = "dr_cr"
    elif exp_hdr or amount_hdr:
        fmt = "amount"
    else:
        return {"kind": "undeposited", "tab": tab, "tab_month": tab_month,
                "statement_format": "unknown", "line_count": 0, "lines": [],
                "issues": ["no recognizable header"], "totals": {}}, \
               ["no header"]

    lines = []
    issues = []
    totals = {}
    prev_day = None

    def _dp(cell):
        nonlocal prev_day
        day, kind = _decode_day_monthly(cell, tab_month, prev_day)
        if day is not None:
            prev_day = day
        if tab_month and day:
            try:
                return datetime.date(2025, tab_month, day).isoformat(), kind
            except ValueError:
                return None, kind
        return None, kind

    def _raw(cell):
        return (cell.isoformat() if isinstance(cell, datetime.datetime)
                else (cell if cell is not None else ""))

    seq = 0

    def _emit(**kw):
        nonlocal seq
        seq += 1
        base = {"sequence": seq * 10, "date_raw": "", "date_parsed": None,
                "details": "", "acc_code": "", "section": "statement",
                "invoice_no": "", "debit": None, "credit": None,
                "amount": None, "currency": "USD", "note": ""}
        base.update(kw)
        lines.append(base)

    if fmt == "two_table":
        r_idx, r_off = receipts_hdr[0]
        # detect a ZWG second amount column from the super-header above
        zwg_col = None
        if r_idx >= 1:
            for j, x in enumerate(rows[r_idx - 1]):
                if _norm(x) == "zwg":
                    zwg_col = j
        # receipts: from r_idx+1 until expenses header (or end gap)
        exp_at = exp_hdr[0][0] if exp_hdr else len(rows)
        for i in range(r_idx + 1, exp_at):
            row = list(rows[i])
            if not any(x is not None for x in row):
                continue
            dl = _norm(row[r_off + 1] if len(row) > r_off + 1 else None)
            if any(s in dl for s in _SUMMARY):
                continue
            # totals row = a lone-ish numeric row with no details
            nums = [x for x in row if isinstance(x, (int, float))]
            if (row[r_off + 1] is None) and nums:
                totals.setdefault("receipts", nums)
                continue
            date_cell = row[r_off] if len(row) > r_off else None
            details = row[r_off + 1] if len(row) > r_off + 1 else None
            inv_no = row[r_off + 2] if len(row) > r_off + 2 else None
            usd = _num(row[r_off + 3]) if len(row) > r_off + 3 else None
            zwg = _num(row[zwg_col]) if (zwg_col is not None
                                         and len(row) > zwg_col) else None
            # remaining non-numeric trailing cells -> note (method / '450??')
            extras = [str(x).strip() for k, x in enumerate(row)
                      if k > r_off + 3 and isinstance(x, str) and x.strip()
                      and k != zwg_col]
            dp, _k = _dp(date_cell)
            if zwg is not None and (usd is None):
                _emit(date_raw=_raw(date_cell), date_parsed=dp, section="receipt",
                      details=(str(details).strip() if details else ""),
                      invoice_no=(str(inv_no).strip() if isinstance(inv_no, str)
                                  else ""),
                      amount=zwg, currency="ZWG", note="; ".join(extras))
            else:
                _emit(date_raw=_raw(date_cell), date_parsed=dp, section="receipt",
                      details=(str(details).strip() if details else ""),
                      invoice_no=(str(inv_no).strip() if isinstance(inv_no, str)
                                  else (str(inv_no) if inv_no is not None
                                        else "")),
                      amount=usd, currency="USD",
                      note=("; ".join(extras)
                            + ((" | ZWG=%s" % zwg) if zwg is not None else "")))
        # expenses table
        if exp_hdr:
            e_idx, e_off = exp_hdr[0]
            for i in range(e_idx + 1, len(rows)):
                row = list(rows[i])
                if not any(x is not None for x in row):
                    continue
                details = row[e_off + 1] if len(row) > e_off + 1 else None
                dl = _norm(details)
                if any(s in dl for s in _SUMMARY):
                    continue
                nums = [x for x in row if isinstance(x, (int, float))]
                if details is None and nums:
                    totals.setdefault("expenses", nums)
                    continue
                date_cell = row[e_off] if len(row) > e_off else None
                amount = _num(row[e_off + 2]) if len(row) > e_off + 2 else None
                acc = row[e_off + 3] if len(row) > e_off + 3 else None
                dp, _k = _dp(date_cell)
                _emit(date_raw=_raw(date_cell), date_parsed=dp,
                      section="expense",
                      details=(str(details).strip() if details else ""),
                      acc_code=(str(acc).strip() if acc else ""),
                      amount=amount, currency="USD")
    elif fmt == "dr_cr":
        h_idx, h_off = drcr_hdr[0]
        for i in range(h_idx + 1, len(rows)):
            row = list(rows[i])
            if not any(x is not None for x in row):
                continue
            details = row[h_off + 1] if len(row) > h_off + 1 else None
            dl = _norm(details)
            if any(s in dl for s in _SUMMARY):
                continue
            if details is None:
                nums = [x for x in row if isinstance(x, (int, float))]
                if nums:
                    totals.setdefault("statement", nums)
                continue
            date_cell = row[h_off]
            dr = _num(row[h_off + 3]) if len(row) > h_off + 3 else None
            cr = _num(row[h_off + 4]) if len(row) > h_off + 4 else None
            dp, _k = _dp(date_cell)
            _emit(date_raw=_raw(date_cell), date_parsed=dp, section="statement",
                  details=(str(details).strip() if details else ""),
                  debit=dr, credit=cr,
                  amount=(cr if cr is not None else dr))
    else:  # amount
        h = (amount_hdr or exp_hdr)[0]
        h_idx, h_off = h
        for i in range(h_idx + 1, len(rows)):
            row = list(rows[i])
            if not any(x is not None for x in row):
                continue
            details = row[h_off + 1] if len(row) > h_off + 1 else None
            dl = _norm(details)
            if any(s in dl for s in _SUMMARY):
                continue
            if details is None:
                nums = [x for x in row if isinstance(x, (int, float))]
                if nums:
                    totals.setdefault("statement", nums)
                continue
            date_cell = row[h_off]
            acc = row[h_off + 2] if len(row) > h_off + 2 else None
            amount = _num(row[h_off + 3]) if len(row) > h_off + 3 else None
            # 'amount' header is at index +3 when Acc Code present; some tabs
            # put Amount at +2 (no Acc col). Fall back.
            if amount is None and isinstance(row[h_off + 2], (int, float)):
                amount = _num(row[h_off + 2])
                acc = None
            dp, _k = _dp(date_cell)
            _emit(date_raw=_raw(date_cell), date_parsed=dp, section="statement",
                  details=(str(details).strip() if details else ""),
                  acc_code=(str(acc).strip() if acc else ""), amount=amount)

    sum_amt = round(sum((l["amount"] or 0.0) for l in lines), 2)
    return {
        "kind": "undeposited", "tab": tab, "tab_month": tab_month,
        "name": "Undeposited %s 2025" % (calendar.month_name[tab_month]
                                         if tab_month else tab),
        "period_month": ("2025-%02d-01" % tab_month) if tab_month else None,
        "statement_format": fmt, "currency_default": "USD",
        "line_count": len(lines), "lines": lines,
        "totals": totals, "v_sum_amount": sum_amt,
        "v_zwg_lines": sum(1 for l in lines if l["currency"] == "ZWG"),
        "v_sections": {s: sum(1 for l in lines if l["section"] == s)
                       for s in ("receipt", "expense", "statement")},
        "issues": issues,
    }, issues


SUSPENSE_TABS = {2025: " Suspense Account", 2026: "2026 Suspense Acc"}
UNDEP_TABS_2025 = ["January Undeposited funds", "February Undeposited funds",
                   "Undeposited April ", "Undeposited May",
                   "Undeposited June ", "July Undeposited"]


def main():
    import openpyxl
    out = None
    args = [a for a in sys.argv[1:]]
    if args and args[-1].endswith(".json"):
        out = args[-1]
    wb25 = openpyxl.load_workbook(DEFAULT_2025, data_only=True)
    wb26 = openpyxl.load_workbook(DEFAULT_2026, data_only=True)
    suspense = []
    for yr, tab in SUSPENSE_TABS.items():
        wb = wb25 if yr == 2025 else wb26
        if tab in wb.sheetnames:
            rows = list(wb[tab].iter_rows(values_only=True))
            stmt, _e = parse_suspense(rows, tab, yr)
            if stmt:
                suspense.append(stmt)
    undeposited = []
    skipped = []
    for tab in UNDEP_TABS_2025:
        if tab not in wb25.sheetnames:
            skipped.append({"tab": tab, "why": "missing"})
            continue
        rows = list(wb25[tab].iter_rows(values_only=True))
        stmt, _e = parse_undeposited(rows, tab)
        undeposited.append(stmt)
    payload = {"suspense": suspense, "undeposited": undeposited,
               "skipped": skipped}
    if out:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=1, default=str)
    return payload


if __name__ == "__main__":
    p = main()
    print("SUSPENSE (%d):" % len(p["suspense"]))
    for s in p["suspense"]:
        print("  %s  lines=%d closing=%s  sum_dr=%s sum_cr=%s  balance_eq=%s "
              "recon_fails=%d  issues=%s"
              % (s["name"], s["line_count"], s["closing_balance"],
                 s["v_sum_dr"], s["v_sum_cr"], s["v_balance_eq"],
                 len(s["v_recon_fails"]), s["issues"]))
    print("\nUNDEPOSITED (%d):" % len(p["undeposited"]))
    for s in p["undeposited"]:
        print("  %-26s fmt=%-9s lines=%-3d sections=%s zwg=%s sum_amt=%s "
              "totals=%s"
              % (s["tab"], s["statement_format"], s["line_count"],
                 s.get("v_sections"), s.get("v_zwg_lines"),
                 s.get("v_sum_amount"), s.get("totals")))
