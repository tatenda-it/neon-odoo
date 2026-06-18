# -*- coding: utf-8 -*-
"""Wages-sheet crew roster -> JSON (LOCAL, reference-only).

Scans the Technician column across ALL weekly sheets (both the old
WEEK|TECHNICIAN layout and the new wide 'Column 1' layout) and maps every raw
spelling to its canonical person using the de-dup map Tatenda confirmed at the
2026-06-18 gate. Every raw spelling is preserved as an alias. Any raw value not
in the curated map is SURFACED (never silently dropped).

Usage: python parse_crew.py [out.json]
"""
import collections
import json
import re
import sys

DEFAULT_XLSX = r"C:\Users\Neon\Downloads\Wages Spread Sheet.xlsx"
STOP = {"total", "totals", "paid", "week", "technician", "jobs covered",
        "grand total", "amount", "column 1", "column 2", "column 3",
        "column 4", "bonus", "name", "", "tbc"}

# Gate-resolved de-dup (CASED raw spelling -> canonical real name). These cased
# keys ARE the known aliases per person (so every member has its aliases even if
# a rare spelling isn't re-found in the scan); the scan validates + adds any
# case variants and surfaces anything unaccounted-for.
CURATED = {
    "Oswell": "Oswel Kauni", "Oswel Kauni": "Oswel Kauni",
    "Stanley": "Stanley Phiri", "Stanely": "Stanley Phiri",
    "Stanely Phiri": "Stanley Phiri",
    "Trymore": "Trymore Mukungu", "Trymore Mukungu": "Trymore Mukungu",
    "Bothwell": "Bothwell Kaposora", "Bothwell Kaposora": "Bothwell Kaposora",
    "Biriad": "Kudzai Mushore", "Kudzai": "Kudzai Mushore",
    "Kudzai Mushore": "Kudzai Mushore",
    "John": "John Gatsi", "John Gatsi": "John Gatsi",
    "Tadiwa": "Tadiwa Marisa", "Tadiwa Marisa": "Tadiwa Marisa",
    "Kelvin": "Kelvin Maibeki", "Kelvin Maibeki": "Kelvin Maibeki",
    "Kevin": "Kelvin Maibeki",
    "Anyway": "Anyway Nyaguwo", "Anyway Nyaguwo": "Anyway Nyaguwo",
    "Adam": "Adam Manwere", "Adam Manwere": "Adam Manwere",
    "Anorld": "Arnold Mutasa", "Arnold": "Arnold Mutasa",
    "Arnold Mutasa": "Arnold Mutasa",
    "Danny": "Kelvin Mushore",                 # DISTINCT person (not Maibeki)
    "KK": "Ranganai",                          # Ranganai's nickname (lead)
    # former crew (kept inactive)
    "Romo": "Romo", "Simba": "Simba", "Hailey": "Hailey", "Brian": "Brian",
    "Ricky": "Ricky", "Doubt": "Doubt", "Steven": "Steven",
    "Tawanda": "Tawanda", "Kudzai Nyanguwo": "Kudzai Nyanguwo",
}
_NCUR = {}  # normalized spelling -> canonical (for scan validation)
ACTIVE = {"Oswel Kauni", "Stanley Phiri", "Trymore Mukungu",
          "Bothwell Kaposora", "Kudzai Mushore", "John Gatsi", "Tadiwa Marisa",
          "Kelvin Maibeki", "Anyway Nyaguwo", "Adam Manwere", "Arnold Mutasa",
          "Kelvin Mushore", "Ranganai"}
FORMER = {"Romo", "Simba", "Hailey", "Brian", "Ricky", "Doubt", "Steven",
          "Tawanda", "Kudzai Nyanguwo"}


def _norm(s):
    return re.sub(r"\s+", " ", re.sub(r"\s*\(.*?\)", "", s).strip()).lower()


def _plausible(s):
    s2 = re.sub(r"\s*\(.*?\)", "", s).strip()
    if not s2 or s2.lower() in STOP or "$" in s2 or "?" in s2:
        return False
    if re.search(r"\d", s2) or len(s2.split()) > 3:
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z .'-]*$", s2))


def parse(path):
    import openpyxl
    if not _NCUR:
        for k, v in CURATED.items():
            _NCUR[_norm(k)] = v
    wb = openpyxl.load_workbook(path, data_only=True)
    raw = collections.Counter()
    for n in wb.sheetnames:
        rows = list(wb[n].iter_rows(values_only=True))
        # technician column(s): the header-identified one PLUS col 0 (wide
        # sheets) — union catches rare names the header-only scan misses.
        cols = {0}
        for row in rows[:5]:
            for ci, c in enumerate(row):
                if isinstance(c, str) and c.strip().lower() in ("technician",
                                                               "column 1"):
                    cols.add(ci)
        for row in rows:
            for ci in cols:
                v = row[ci] if ci < len(row) else None
                if isinstance(v, str) and _plausible(v):
                    raw[re.sub(r"\s*\(.*?\)", "", v).strip()] += 1

    members = {}
    for c in (ACTIVE | FORMER):
        is_lead = (c == "Ranganai")
        status = "active" if c in ACTIVE else "former"
        # seed aliases from the curated cased keys (guarantees every member has
        # its known spellings even if a rare one isn't re-found in the scan).
        members[c] = {
            "name": c,
            "aliases": set(k for k, vv in CURATED.items() if vv == c),
            "role": "lead" if is_lead else "unknown",
            "is_lead": is_lead, "status": status,
            "active": status == "active", "source": "wages_sheet",
        }
    unmapped = collections.Counter()
    for rawn, cnt in raw.items():
        canon = _NCUR.get(_norm(rawn))
        if canon:
            members[canon]["aliases"].add(rawn)  # add the scanned spelling too
        else:
            unmapped[rawn] += cnt

    out_members = []
    for c in sorted(members, key=lambda k: (not members[k]["is_lead"],
                                            members[k]["status"], k)):
        m = members[c]
        m2 = dict(m)
        m2["aliases"] = "\n".join(sorted(m["aliases"]))
        m2["alias_list"] = sorted(m["aliases"])
        out_members.append(m2)
    return {"members": out_members, "unmapped": dict(unmapped)}


def main():
    out = next((a for a in sys.argv[1:] if a.endswith(".json")), None)
    payload = parse(DEFAULT_XLSX)
    if out:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=1, default=str)
    return payload


if __name__ == "__main__":
    p = main()
    act = [m for m in p["members"] if m["status"] == "active"]
    fmr = [m for m in p["members"] if m["status"] == "former"]
    print("CREW ROSTER: %d members (%d active, %d former)"
          % (len(p["members"]), len(act), len(fmr)))
    for m in p["members"]:
        print("  [%-6s] %-20s lead=%-5s aliases=%s"
              % (m["status"], m["name"], m["is_lead"], m["alias_list"]))
    print("UNMAPPED (surfaced, none expected):", p["unmapped"])
