# -*- coding: utf-8 -*-
"""Petty-cash cashbook xlsx -> JSON parser (LOCAL, reference-only).

Reads "Neon Expenses 2026.xlsx", parses each "* Petty Cash" month tab VERBATIM
into JSON for the prod loader (mirrors the Zoho JSON->loader flow). Stores raw
cell values; the verify checks are ASSERTIONS, not corrections.

DATE DECODE (the cashbook stores the day in a mangled way):
  * day 1-12 -> Excel datetime(YEAR, DAY, TAB_MONTH): the MIDDLE field is the
    day-of-month; the LAST field is the tab month. So recovered_day = cell.month,
    and cell.day MUST equal the tab month (sanity).
  * day 13-31 -> text "DD-MM-YY": recovered_day = DD; MM must equal tab month.
  * blank -> carry forward the previous row's recovered day.
  YEAR + MONTH come from the tab's own period-header datetime (NOT a constant),
  cross-checked against the tab-name month. date_parsed = date(tab_year,
  tab_month, recovered_day). date_raw (original cell) is ALWAYS preserved.

Usage:  python addons/neon_migration/scripts/parse_petty_cash.py [xlsx] [out.json]
"""
import calendar
import datetime
import json
import re
import sys

# NB: openpyxl is imported lazily inside main() so the pure parsing helpers
# (parse_tab / _decode_day / _find_header ...) can be unit-tested in an
# environment without openpyxl (e.g. the Odoo shell), passing synthetic rows.

DEFAULT_XLSX = r"C:\Users\Neon\Downloads\Neon Expenses 2026.xlsx"
_MONTH_WORDS = {}
for _i, _full in enumerate(calendar.month_name):
    if _full:
        _MONTH_WORDS[_full.lower()] = _i
for _i, _ab in enumerate(calendar.month_abbr):
    if _ab:
        _MONTH_WORDS[_ab.lower()] = _i
_MONTH_WORDS["sept"] = 9  # common non-standard abbreviation
_SUMMARY_LABELS = ("closing balance", "opening balance carried", "total",
                   "totals", "grand total")


def _norm(c):
    return str(c).strip().lower() if c is not None else ""


def _month_from_tabname(tab):
    """Scan ALL alpha words for a month name/abbr (2025 tabs put the month
    last, e.g. 'Petty Cash May'; 2026 put it first, e.g. 'January Petty
    Cash'). Returns the month number, or None."""
    for w in re.findall(r"[a-zA-Z]+", (tab or "").lower()):
        if w in _MONTH_WORDS:
            return _MONTH_WORDS[w]
    return None


def _find_header(rows):
    """Return (row_index, col_offset). The table does NOT always start in
    column A -- October 2025 is shifted one column right. Scan every possible
    start column for the Date/Details/.../Balance header."""
    for i, row in enumerate(rows):
        rl = list(row)
        for c in range(0, max(1, len(rl) - 5)):
            seg = [_norm(x) for x in (rl[c:c + 6] + [""] * 6)[:6]]
            if seg[0].startswith("date") and seg[1].startswith("details") \
                    and seg[5].startswith("balance"):
                return i, c
    return None, None


def _find_period_dt(rows, header_idx):
    for i in range(header_idx - 1, -1, -1):
        for x in rows[i]:
            if isinstance(x, datetime.datetime):
                return x
    return None


def _find_closing(rows, header_idx):
    for i in range(header_idx - 1, -1, -1):
        row = list(rows[i])
        for j in range(len(row)):
            if isinstance(row[j], str) and row[j].strip().lower() == "balance":
                for k in range(j + 1, len(row)):
                    if isinstance(row[k], (int, float)):
                        return float(row[k])
    return None


def _decode_day(cell, tab_month, prev_day):
    """Recover day-of-month. Two Excel encodings coexist across tabs:
      * MANGLED: datetime(Y, DAY, TAB_MONTH) -- the month-field holds the day,
        the day-field holds the tab month. Detect: cell.day == tab_month ->
        day = cell.month.
      * REAL: datetime(Y, TAB_MONTH, DAY) -- ordinary date. Detect:
        cell.month == tab_month -> day = cell.day.
      * text "DD-MM-YY" -> day = DD (sanity MM == tab_month).
      * blank / unparseable note -> carry the previous row's day.
    Returns (day, kind, sanity_ok)."""
    if cell is None or (isinstance(cell, str) and not cell.strip()):
        return prev_day, "carry", None
    if isinstance(cell, datetime.datetime):
        if cell.day == tab_month:
            return cell.month, "dt-mangled", True
        if cell.month == tab_month:
            return cell.day, "dt-real", True
        return prev_day, "dt-ambiguous", False
    if isinstance(cell, str):
        m = re.match(r"\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{2,4})",
                     cell.strip())
        if m:
            return int(m.group(1)), "text", (int(m.group(2)) == tab_month)
        return prev_day, "text-note", None  # free-text in the date cell
    return prev_day, "unparsed", False


def parse_tab(rows, tab):
    """Parse one tab's rows (a list of cell-tuples) into a statement dict.
    openpyxl-free so it can be unit-tested with synthetic rows."""
    header_idx, col0 = _find_header(rows)
    issues = []
    if header_idx is None:
        return None, ["no header row found"]
    if col0:
        issues.append("table shifted to column offset %d" % col0)
    period_dt = _find_period_dt(rows, header_idx)
    name_month = _month_from_tabname(tab)
    # MONTH is authoritative from the TAB NAME (the Dec-2025 header datetime is
    # mis-entered as Nov; the name is right). YEAR comes from the header
    # datetime (tab names carry no year).
    tab_month = name_month or (period_dt.month if period_dt else None)
    tab_year = period_dt.year if period_dt else None
    if period_dt and name_month and period_dt.month != name_month:
        issues.append("header month %s != tab-name %s -> using tab-name"
                      % (period_dt.month, name_month))
    if tab_year is None:
        issues.append("no period datetime -> YEAR UNKNOWN")
    if tab_month is None:
        issues.append("no month resolvable from tab name or header")
    closing_header = _find_closing(rows, header_idx)

    # Collect transactions (rows with non-empty details), capture cr_total
    # (first lone numeric after the last transaction).
    txns = []
    cr_total = None
    gap = 0
    i = header_idx + 1
    while i < len(rows):
        row = list(rows[i][col0:col0 + 6]) + [None] * 6
        details = row[1]
        dl = str(details).strip().lower() if details else ""
        has_any = any(x is not None for x in row[:6])
        is_summary = any(lbl in dl for lbl in _SUMMARY_LABELS) if dl else False
        if details is not None and dl and not is_summary:
            txns.append((i + 1, row))  # 1-indexed sheet row
            gap = 0
        else:
            if is_summary:
                # "Closing Balance" / "Total" row: the Cr column carries the
                # period Cr-total -> capture for cross-check; NEVER counted as
                # a transaction.
                crc = row[4]
                if isinstance(crc, (int, float)) and cr_total is None:
                    cr_total = float(crc)
            elif has_any:
                nums = [x for x in row[:6] if isinstance(x, (int, float))]
                if txns and cr_total is None and len(nums) == 1:
                    cr_total = float(nums[0])
            gap += 1
            if txns and gap >= 8:
                break
        i += 1

    lines = []
    prev_day = None
    prev_balance = None
    sum_dr = sum_cr = 0.0
    mono_ok = True
    sanity_fail = []
    recon_fail = []
    for seq, (sheet_row, row) in enumerate(txns, start=1):
        date_cell, details, acc, dr, cr, bal = row[:6]
        day, kind, sane = _decode_day(date_cell, tab_month, prev_day)
        if sane is False:
            sanity_fail.append((sheet_row, repr(date_cell)))
        if day is not None and prev_day is not None and day < prev_day:
            mono_ok = False
        if day is not None:
            prev_day = day
        dr = float(dr) if isinstance(dr, (int, float)) else 0.0
        cr = float(cr) if isinstance(cr, (int, float)) else 0.0
        bal = float(bal) if isinstance(bal, (int, float)) else None
        sum_dr += dr
        sum_cr += cr
        # row-to-row reconcile (skip the opening row where prev_balance None)
        if prev_balance is not None and bal is not None:
            expect = prev_balance + dr - cr
            if abs(expect - bal) > 0.01:
                recon_fail.append((sheet_row, prev_balance, dr, cr, bal,
                                   round(expect, 2)))
        if bal is not None:
            prev_balance = bal
        date_parsed = None
        if tab_year and tab_month and day:
            try:
                date_parsed = datetime.date(tab_year, tab_month, day).isoformat()
            except ValueError:
                date_parsed = None
                issues.append("row %d invalid date y%s m%s d%s"
                              % (sheet_row, tab_year, tab_month, day))
        lines.append({
            "sequence": seq * 10,
            "sheet_row": sheet_row,
            "date_raw": (date_cell.isoformat()
                         if isinstance(date_cell, datetime.datetime)
                         else (date_cell if date_cell is not None else "")),
            "date_parsed": date_parsed,
            "decode_kind": kind,
            "details": (str(details).strip() if details else ""),
            "acc_code": (str(acc).strip() if acc else ""),
            "debit": dr,
            "credit": cr,
            "balance": bal,
        })

    opening_balance = lines[0]["balance"] if lines else None
    closing_line = lines[-1]["balance"] if lines else None
    period_month = (datetime.date(tab_year, tab_month, 1).isoformat()
                    if (tab_year and tab_month) else None)

    stmt = {
        "tab": tab,
        "name": "%s %s" % (calendar.month_name[tab_month] if tab_month else "?",
                           tab_year or "?"),
        "period_month": period_month,
        "tab_year": tab_year,
        "tab_month": tab_month,
        "currency_code": "USD",
        "opening_balance": opening_balance,
        "closing_balance": closing_line,
        "closing_header": closing_header,
        "cr_total": cr_total,
        "source_tab": tab,
        "line_count": len(lines),
        "lines": lines,
        # verify
        "v_monotonic_days": mono_ok,
        "v_sum_dr": round(sum_dr, 2),
        "v_sum_cr": round(sum_cr, 2),
        "v_crtotal_match": (cr_total is not None
                            and abs(cr_total - sum_cr) < 0.01),
        "v_closing_match": (closing_header is not None and closing_line is not None
                            and abs(closing_header - closing_line) < 0.01),
        # The opening row is stored verbatim with its Dr = the opening float, so
        # the whole statement flows from zero: closing == sum(Dr) - sum(Cr).
        "v_balance_eq": (closing_line is not None
                         and abs((sum_dr - sum_cr) - closing_line) < 0.01),
        "v_sanity_fails": sanity_fail,
        "v_recon_fails": recon_fail,
        "issues": issues,
    }
    return stmt, issues


def main():
    import openpyxl  # lazy — only the CLI/loader path needs it
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out = sys.argv[2] if len(sys.argv) > 2 else None
    wb = openpyxl.load_workbook(path, data_only=True)
    pc_tabs = [n for n in wb.sheetnames if "petty cash" in n.lower()]
    statements = []
    skipped = []
    for tab in pc_tabs:
        rows = list(wb[tab].iter_rows(values_only=True))
        stmt, errs = parse_tab(rows, tab)
        if stmt:
            statements.append(stmt)
        else:
            skipped.append({"tab": tab, "errors": errs})
    payload = {"source_file": path, "pc_tab_count": len(pc_tabs),
               "statement_count": len(statements), "skipped": skipped,
               "statements": statements}
    if out:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=1, default=str)
    return payload


if __name__ == "__main__":
    p = main()
    # Compact report to stdout.
    print("PETTY-CASH PARSE — %d/%d petty-cash tabs parsed; skipped=%s"
          % (p["statement_count"], p["pc_tab_count"],
             [s["tab"] for s in p["skipped"]] or "none"))
    for s in p["statements"]:
        print("\n" + "=" * 70)
        print("TAB %r -> %s  period_month=%s  (year=%s month=%s)"
              % (s["tab"], s["name"], s["period_month"],
                 s["tab_year"], s["tab_month"]))
        print("  lines=%d opening=%s closing=%s closing_header=%s cr_total=%s"
              % (s["line_count"], s["opening_balance"], s["closing_balance"],
                 s["closing_header"], s["cr_total"]))
        print("  VERIFY: monotonic_days=%s  sum_dr=%s sum_cr=%s "
              "crtotal_match=%s  closing_match=%s  balance_eq=%s  "
              "recon_fails=%d sanity_fails=%d"
              % (s["v_monotonic_days"], s["v_sum_dr"], s["v_sum_cr"],
                 s["v_crtotal_match"], s["v_closing_match"], s["v_balance_eq"],
                 len(s["v_recon_fails"]), len(s["v_sanity_fails"])))
        if s["v_sanity_fails"]:
            print("  !! date-sanity fails:", s["v_sanity_fails"][:5])
        if s["v_recon_fails"]:
            print("  !! balance-recon fails:", s["v_recon_fails"][:5])
        if s["issues"]:
            print("  !! issues:", s["issues"])
