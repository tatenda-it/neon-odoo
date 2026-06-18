# -*- coding: utf-8 -*-
"""Wages sheet -> JSON (LOCAL, reference-only). WEEKLY-LUMP pay per technician.

Handles all three layouts in "Wages Spread Sheet.xlsx":
  A) multi-week-per-sheet: WEEK | TECHNICIAN | JOBS COVERED | TOTAL  (week label
     rows in col0, a 'TOTAL' row per week)
  B) single-week, tech in col0: TECHNICIAN | JOBS COVERED | TOTAL | Paid
  C) wide single-week: Column 1 | JOBS COVERED | TOTAL
Pay = the TOTAL column (the weekly lump; NO per-job split). week_label is kept
VERBATIM; week_date is best-effort (None if unparseable, never fabricated).
Crew-FK + job links are resolved by the LOADER (need prod data); this only
parses + reconciles (sum-of-techs == the week TOTAL row).

Usage: python parse_wages.py [out.json]
"""
import calendar
import collections
import datetime
import json
import re
import sys

DEFAULT_XLSX = r"C:\Users\Neon\Downloads\Wages Spread Sheet.xlsx"
STOP = {"total", "totals", "grand total", "paid", "week", "technician",
        "jobs covered", "amount", "column 1", "column 2", "column 3",
        "column 4", "bonus", "name", "", "tbc"}
_MON = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
_MON.update({m.lower(): i for i, m in enumerate(calendar.month_abbr) if m})


def _norm(c):
    return str(c).strip().lower() if isinstance(c, str) else ""


def _num(c):
    return float(c) if isinstance(c, (int, float)) else None


def _plausible_name(s):
    s2 = re.sub(r"\s*\(.*?\)", "", s).strip()
    if not s2 or s2.lower() in STOP or "$" in s2 or "?" in s2:
        return False
    if re.search(r"\d", s2) or len(s2.split()) > 3:
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z .'-]*$", s2))


def _date_from_text(s, default_year=None):
    """Best-effort date from a label/sheet-name. Returns ISO or None."""
    if not s:
        return None
    s = str(s)
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)  # dd/mm/yy
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + y if y < 100 else y
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return None
    yr = None
    my = re.search(r"(20\d\d)", s)
    if my:
        yr = int(my.group(1))
    mon = next((v for k, v in _MON.items()
                if re.search(r"\b%s" % re.escape(k), s.lower())), None)
    dm = re.search(r"\b(\d{1,2})\b", re.sub(r"20\d\d", "", s))
    if mon and dm:
        y = yr or default_year or 2025
        try:
            return datetime.date(y, mon, int(dm.group(1))).isoformat()
        except ValueError:
            return None
    return None


def _find_header(rows):
    for i, row in enumerate(rows[:6]):
        nm = [_norm(c) for c in row]
        if "jobs covered" in nm and "total" in nm:
            cols = {"jobs": nm.index("jobs covered"), "total": nm.index("total"),
                    "tech": None, "week": None}
            if "technician" in nm:
                cols["tech"] = nm.index("technician")
            elif "column 1" in nm:
                cols["tech"] = nm.index("column 1")
            if "week" in nm:
                cols["week"] = nm.index("week")
            cols["paid"] = cols["total"] + 1
            return i, cols
    return None, None


def parse_sheet(rows, n):
    """Parse one sheet's rows -> (entries, weeks). openpyxl-free (unit-testable
    with synthetic rows)."""
    entries, weeks = [], []
    hi, cols = _find_header(rows)
    if hi is None:
        return entries, weeks
    tc, jc, vc, wc = cols["tech"], cols["jobs"], cols["total"], cols["week"]
    pc = cols["paid"]
    sheet_year = 2026 if "2026" in n else 2025
    placeholder = "%s (unlabeled week)" % n.strip()

    def new_week(label, date):
        return {"label": label, "date": date, "sum": 0.0, "total_row": None,
                "n": 0, "buf": []}

    st = new_week(n.strip(), _date_from_text(n, sheet_year))

    def _flush():
        if st["n"] or st["total_row"] is not None:
            lbl = st["label"] or placeholder
            weeks.append({"sheet": n, "label": lbl, "date": st["date"],
                          "total_row": st["total_row"],
                          "sum_techs": round(st["sum"], 2)})
            for e in st["buf"]:           # stamp final label/date onto entries
                e["week_label"] = lbl
                e["week_date"] = st["date"]
                entries.append(e)

    for r in rows[hi + 1:]:
        row = list(r)
        tv = row[tc] if (tc is not None and tc < len(row)) else None
        wv = row[wc] if (wc is not None and wc < len(row)) else None
        total = _num(row[vc]) if vc < len(row) else None
        jobs = row[jc] if jc < len(row) else None
        paid = row[pc] if pc < len(row) else None
        tvs = tv.strip() if isinstance(tv, str) else ""
        # TOTAL row = week boundary (its label sits in col0 in the multi-week
        # layout; some later weeks are separated ONLY by a TOTAL row).
        is_total_label = tvs.lower() in ("total", "totals", "grand total") \
            or _norm(row[0] if row else "") in ("total", "totals",
                                                "grand total")
        if is_total_label or (not tvs and total is not None and not jobs):
            if total is not None:
                st["total_row"] = total
            _flush()
            st = new_week(None, None)     # next week unlabeled until a label row
            continue
        # week-label row: relabel an empty week, else flush + start a new one
        if wc is not None and isinstance(wv, str) and wv.strip() and not tvs:
            if st["n"] == 0 and st["total_row"] is None:
                st["label"] = wv.strip()
                st["date"] = _date_from_text(wv, sheet_year)
            else:
                _flush()
                st = new_week(wv.strip(), _date_from_text(wv, sheet_year))
            continue
        if tvs and _plausible_name(tvs):
            tval = total or 0.0
            st["sum"] += tval
            st["n"] += 1
            paid_s = "paid" if (isinstance(paid, str)
                                and "paid" in paid.lower()) else "unknown"
            st["buf"].append({
                "sheet": n, "week_label": st["label"], "week_date": st["date"],
                "technician_raw": re.sub(r"\s*\(.*?\)", "", tvs).strip(),
                "total": tval, "currency_code": "USD", "paid": paid_s,
                "jobs_raw": (str(jobs).strip() if jobs is not None else ""),
                "source": "wages_sheet",
            })
    _flush()
    return entries, weeks


def parse(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    entries = []
    weeks = []        # {sheet,label,date,total_row,sum_techs}
    noncurrency = []
    for n in wb.sheetnames:
        e, w = parse_sheet(list(wb[n].iter_rows(values_only=True)), n)
        entries.extend(e)
        weeks.extend(w)

    # reconciliation
    recon_fail = []
    for w in weeks:
        if w["total_row"] is not None \
                and abs(w["total_row"] - w["sum_techs"]) > 0.01:
            recon_fail.append(w)
    return {"entries": entries, "weeks": weeks, "recon_fail": recon_fail,
            "noncurrency": noncurrency}


def main():
    out = next((a for a in sys.argv[1:] if a.endswith(".json")), None)
    p = parse(DEFAULT_XLSX)
    if out:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(p, f, indent=1, default=str)
    return p


if __name__ == "__main__":
    p = main()
    e = p["entries"]
    dates = [x["week_date"] for x in e if x["week_date"]]
    print("WAGES PARSE: %d entries, %d weeks, recon_fails=%d"
          % (len(e), len(p["weeks"]), len(p["recon_fail"])))
    print("  total $: %.2f | week_date span: %s -> %s | null week_date: %d"
          % (sum(x["total"] for x in e), min(dates) if dates else "-",
             max(dates) if dates else "-", sum(1 for x in e if not x["week_date"])))
    print("  distinct technician_raw: %d"
          % len(set(x["technician_raw"] for x in e)))
    print("\n-- sample entries (each layout) --")
    seen = set()
    for x in e:
        key = "multi" if "week" in x["sheet"].lower() else "single"
        if key not in seen:
            seen.add(key)
            print("  [%s] %s | %s | $%s | paid=%s | jobs=%r"
                  % (key, x["week_label"], x["technician_raw"], x["total"],
                     x["paid"], x["jobs_raw"][:50]))
    print("\n-- recon fails (first 8) --")
    for w in p["recon_fail"][:8]:
        print("  %s: sum_techs=%s total_row=%s" % (w["label"], w["sum_techs"],
                                                   w["total_row"]))
